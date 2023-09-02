from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
import globalConstants

class Test_Demo:

    # Her test öncesi çağırılır
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        # Günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur 
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
    # Her test sonrası çağırılır
    def teardown_method(self):
        self.driver.quit()
    # setup -> test_demo -> teardown
    def test_demo(self):
        text = "Hello"
        assert text == "Hello"
    # setup -> test_demo2 -> teardown
    def test_demo2(self):
        assert True
    
    def getData():
        #veri al
        exelFile = openpyxl.load_workbook("data/invalit_login.xlsx")
        selectedSheet = exelFile["Sayfa1"]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data
    #@pytest.mark.skip()
    @pytest.mark.parametrize("username,password",getData())
    def test_invalit_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID , "user-name")
        self.waitForElementVisible((By.ID,"password"))
        passwordInput = self.driver.find_element(By.ID , "password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        errorMesage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalit_login{username}_{password}.png")
        assert errorMesage.text == "Epic sadface: Username and password do not match any user in this service"

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(expected_conditions.visibility_of_element_located((locator)))